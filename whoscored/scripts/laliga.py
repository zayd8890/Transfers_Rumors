from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import platform
import subprocess
from datetime import datetime

# La Liga teams with WhoScored IDs
TEAMS = {
    "Alaves": 1108,
    "Almeria": 330,
    "Athletic Club": 167,
    "Atletico Madrid": 13,
    "Barcelona": 65,
    "Betis": 150,
    "Cadiz": 1421,
    "Celta Vigo": 294,
    "Getafe": 3709,
    "Girona": 12321,
    "Granada": 16795,
    "Las Palmas": 472,
    "Mallorca": 106,
    "Osasuna": 331,
    "Rayo Vallecano": 4726,
    "Real Madrid": 52,
    "Real Sociedad": 166,
    "Sevilla": 67,
    "Valencia": 104,
    "Villarreal": 153
}

# Exact header structure from your screenshot
HEADERS = [
    "Player", "CM", "KG", "Apps", "Mins", "Goals", "Assists", 
    "Yel", "Red", "SpG", "PS%", "AerialsWon", "MotM", "Rating"
]

COL_MAP = [0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]  # TD positions

def initialize_workbook():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"la_liga_players_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    return wb, filename

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_argument("--log-level=3")
    return webdriver.Chrome(options=options)

def scrape_team(driver, team_name, team_id, wb):
    print(f"Scraping {team_name}...")
    try:
        driver.get(f'https://www.whoscored.com/Teams/{team_id}/Show/Spain-{team_name.replace(" ", "-")}')
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, 'top-player-stats-summary-grid')))
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        table = soup.find('table', {'id': 'top-player-stats-summary-grid'})
        
        if not table:
            print(f"No table found for {team_name}")
            return False
        
        # Create worksheet for team
        ws = wb.create_sheet(title=team_name[:31])  # Excel sheet name limit
        
        # Write headers
        ws.append(HEADERS)
        
        # Style headers
        header_fill = PatternFill(start_color='D40026', end_color='D40026', fill_type='solid')  # La Liga red
        for cell in ws[1]:
            cell.fill = header_fill
        
        # Player data
        for tr in table.find_all('tr')[1:]:  # Skip header row
            cols = tr.find_all('td')
            if len(cols) < 14:
                continue
                
            row_data = []
            for idx in COL_MAP:
                try:
                    row_data.append(cols[idx].get_text(strip=True).replace('-', '0'))
                except:
                    row_data.append('0')
            
            ws.append(row_data)
        
        # Highlight top performers (rating >= 7.0)
        highlight_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
        for row in ws.iter_rows(min_row=2):
            try:
                if float(row[13].value or 0) >= 7.0:  # Rating column
                    for cell in row:
                        cell.fill = highlight_fill
            except ValueError:
                pass
        
        # Format columns
        col_widths = {
            'A': 25,  # Player
            'B': 5, 'C': 5, 'D': 5, 'E': 7,
            'F': 5, 'G': 7, 'H': 5, 'I': 5,
            'J': 5, 'K': 5, 'L': 10, 'M': 5,
            'N': 6  # Rating
        }
        
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        return True
    
    except Exception as e:
        print(f"Error scraping {team_name}: {str(e)}")
        return False

def main():
    driver = None
    try:
        wb, filename = initialize_workbook()
        driver = setup_driver()
        
        success_count = 0
        for team_name, team_id in TEAMS.items():
            if scrape_team(driver, team_name, team_id, wb):
                success_count += 1
        
        if success_count == 0:
            raise ValueError("No data scraped from any team")
        
        wb.save(filename)
        print(f"\nSuccess! Saved data for {success_count} teams to {filename}")
        
        # Auto-open file
        filepath = os.path.abspath(filename)
        if platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Darwin':
            subprocess.run(['open', filepath])
        else:
            subprocess.run(['xdg-open', filepath])
            
    except Exception as e:
        print(f"\nFatal error: {str(e)}")
    finally:
        if driver:
            driver.quit()
        if 'wb' in locals():
            wb.close()

if __name__ == "__main__":
    main()